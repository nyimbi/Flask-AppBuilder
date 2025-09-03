"""
Excel Export Implementation.

Provides comprehensive Excel export functionality with formatting, styling,
charts, and multiple worksheet support.
"""

import logging
from io import BytesIO
from typing import List, Dict, Any, Optional, Union
from datetime import datetime, date
from decimal import Decimal

log = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    log.warning("openpyxl not available - Excel export functionality limited")


class ExcelExporter:
    """
    Excel format exporter with advanced formatting and chart capabilities.
    
    Supports multiple worksheets, styling, data validation, and basic charts.
    Requires openpyxl package for full functionality.
    """
    
    def __init__(self):
        """Initialize Excel exporter with default settings."""
        self.default_options = {
            'sheet_name': 'Data',
            'include_metadata': True,
            'include_headers': True,
            'auto_filter': True,
            'freeze_headers': True,
            'date_format': 'yyyy-mm-dd',
            'datetime_format': 'yyyy-mm-dd hh:mm:ss',
            'number_format': '#,##0.00',
            'header_style': {
                'font': {'bold': True, 'color': 'FFFFFF'},
                'fill': {'color': '4F81BD', 'type': 'solid'},
                'alignment': {'horizontal': 'center'}
            },
            'data_style': {
                'font': {'size': 10},
                'alignment': {'vertical': 'top'}
            },
            'column_widths': 'auto',
            'add_charts': False,
            'chart_types': ['bar'],  # bar, line, pie
            'metadata_sheet': False
        }
    
    def export(self, 
               data: List[Dict[str, Any]],
               filename: str,
               metadata: Optional[Dict[str, Any]] = None,
               options: Optional[Dict[str, Any]] = None) -> bytes:
        """
        Export data to Excel format.
        
        Args:
            data: List of dictionaries containing the data
            filename: Target filename
            metadata: Export metadata
            options: Excel-specific export options
            
        Returns:
            Excel content as bytes
            
        Raises:
            ImportError: If openpyxl is not available
            Exception: If export fails
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl package is required for Excel export functionality")
        
        try:
            # Merge options with defaults
            export_options = {**self.default_options, **(options or {})}
            
            # Create workbook
            wb = Workbook()
            
            # Create main data worksheet
            ws = wb.active
            ws.title = export_options['sheet_name']
            
            # Write data to worksheet
            if data:
                self._write_excel_data(ws, data, metadata, export_options)
            else:
                # Handle empty data
                ws['A1'] = "No data available for export"
                ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # Create metadata sheet if requested
            if export_options.get('metadata_sheet', False) and metadata:
                self._create_metadata_sheet(wb, metadata)
            
            # Add charts if requested
            if export_options.get('add_charts', False) and data:
                self._add_charts(ws, data, export_options)
            
            # Save to bytes
            output = BytesIO()
            wb.save(output)
            excel_content = output.getvalue()
            output.close()
            
            log.info(f"Excel export completed: {len(data)} records, {len(excel_content)} bytes")
            return excel_content
            
        except Exception as e:
            log.error(f"Excel export failed: {e}")
            raise
    
    def _write_excel_data(self, worksheet, 
                         data: List[Dict[str, Any]], 
                         metadata: Optional[Dict[str, Any]],
                         options: Dict[str, Any]) -> None:
        """
        Write data to Excel worksheet with formatting.
        
        Args:
            worksheet: openpyxl worksheet object
            data: Data to write
            metadata: Export metadata
            options: Export options
        """
        try:
            row_offset = 1
            
            # Add metadata header if requested
            if options.get('include_metadata', True) and metadata:
                row_offset = self._write_metadata_header(worksheet, metadata, options)
                row_offset += 2  # Add some spacing
            
            # Get headers
            headers = []
            if data:
                # Use all unique keys from all rows
                all_keys = set()
                for row in data:
                    all_keys.update(row.keys())
                headers = sorted(list(all_keys))
            
            # Write headers
            if options.get('include_headers', True) and headers:
                for col_idx, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=row_offset, column=col_idx, value=header)
                    self._apply_header_style(cell, options)
                
                header_row = row_offset
                row_offset += 1
            else:
                header_row = None
            
            # Write data rows
            for row_idx, row_data in enumerate(data, row_offset):
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header)
                    formatted_value = self._format_cell_value(value, options)
                    
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=formatted_value)
                    self._apply_data_style(cell, value, options)
            
            # Apply additional formatting
            self._apply_worksheet_formatting(worksheet, headers, header_row, len(data), options)
            
        except Exception as e:
            log.error(f"Error writing Excel data: {e}")
            raise
    
    def _write_metadata_header(self, worksheet, 
                              metadata: Dict[str, Any], 
                              options: Dict[str, Any]) -> int:
        """
        Write metadata header to worksheet.
        
        Args:
            worksheet: Excel worksheet
            metadata: Metadata to write
            options: Export options
            
        Returns:
            Next available row number
        """
        current_row = 1
        
        try:
            # Title
            title_cell = worksheet.cell(row=current_row, column=1, value="Export Information")
            title_cell.font = Font(bold=True, size=14)
            current_row += 1
            
            # Metadata items
            for key, value in metadata.items():
                if value is not None:
                    # Format key nicely
                    display_key = key.replace('_', ' ').title()
                    
                    # Format value
                    if isinstance(value, (datetime, date)):
                        if isinstance(value, datetime):
                            formatted_value = value.strftime('%Y-%m-%d %H:%M:%S')
                        else:
                            formatted_value = value.strftime('%Y-%m-%d')
                    else:
                        formatted_value = str(value)
                    
                    # Write key-value pair
                    worksheet.cell(row=current_row, column=1, value=f"{display_key}:")
                    worksheet.cell(row=current_row, column=2, value=formatted_value)
                    current_row += 1
            
        except Exception as e:
            log.warning(f"Error writing Excel metadata header: {e}")
        
        return current_row
    
    def _format_cell_value(self, value: Any, options: Dict[str, Any]) -> Any:
        """
        Format a single cell value for Excel.
        
        Args:
            value: Value to format
            options: Export options
            
        Returns:
            Formatted value suitable for Excel
        """
        if value is None:
            return ""
        elif isinstance(value, bool):
            return value
        elif isinstance(value, (int, float, Decimal)):
            return float(value) if isinstance(value, Decimal) else value
        elif isinstance(value, (datetime, date)):
            return value  # Excel handles these natively
        elif isinstance(value, (list, tuple)):
            return "; ".join(str(item) for item in value)
        elif isinstance(value, dict):
            return "; ".join(f"{k}: {v}" for k, v in value.items())
        else:
            return str(value)
    
    def _apply_header_style(self, cell, options: Dict[str, Any]) -> None:
        """Apply styling to header cells."""
        header_style = options.get('header_style', {})
        
        # Font styling
        font_options = header_style.get('font', {})
        cell.font = Font(
            bold=font_options.get('bold', True),
            color=font_options.get('color', 'FFFFFF'),
            size=font_options.get('size', 11)
        )
        
        # Fill styling
        fill_options = header_style.get('fill', {})
        if fill_options.get('type') == 'solid':
            cell.fill = PatternFill(
                start_color=fill_options.get('color', '4F81BD'),
                end_color=fill_options.get('color', '4F81BD'),
                fill_type='solid'
            )
        
        # Alignment
        alignment_options = header_style.get('alignment', {})
        cell.alignment = Alignment(
            horizontal=alignment_options.get('horizontal', 'center'),
            vertical=alignment_options.get('vertical', 'center')
        )
    
    def _apply_data_style(self, cell, value: Any, options: Dict[str, Any]) -> None:
        """Apply styling to data cells."""
        data_style = options.get('data_style', {})
        
        # Font styling
        font_options = data_style.get('font', {})
        if font_options:
            cell.font = Font(
                size=font_options.get('size', 10),
                color=font_options.get('color')
            )
        
        # Number formatting
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            cell.number_format = options.get('number_format', '#,##0.00')
        elif isinstance(value, datetime):
            cell.number_format = options.get('datetime_format', 'yyyy-mm-dd hh:mm:ss')
        elif isinstance(value, date):
            cell.number_format = options.get('date_format', 'yyyy-mm-dd')
        
        # Alignment
        alignment_options = data_style.get('alignment', {})
        if alignment_options:
            cell.alignment = Alignment(
                horizontal=alignment_options.get('horizontal'),
                vertical=alignment_options.get('vertical', 'top'),
                wrap_text=alignment_options.get('wrap_text', True)
            )
    
    def _apply_worksheet_formatting(self, worksheet, headers: List[str], 
                                   header_row: Optional[int], data_rows: int, 
                                   options: Dict[str, Any]) -> None:
        """Apply overall worksheet formatting."""
        try:
            # Auto-adjust column widths
            if options.get('column_widths') == 'auto' and headers:
                for col_idx, header in enumerate(headers, 1):
                    column_letter = worksheet.cell(row=1, column=col_idx).column_letter
                    worksheet.column_dimensions[column_letter].width = min(max(len(str(header)) + 2, 10), 50)
            
            # Add auto filter
            if options.get('auto_filter', True) and header_row and data_rows > 0:
                worksheet.auto_filter.ref = f"A{header_row}:{worksheet.cell(row=header_row + data_rows, column=len(headers)).coordinate}"
            
            # Freeze header row
            if options.get('freeze_headers', True) and header_row:
                worksheet.freeze_panes = f"A{header_row + 1}"
            
            # Add borders to data range
            if headers and data_rows > 0:
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in worksheet.iter_rows(
                    min_row=header_row or 1,
                    max_row=(header_row or 1) + data_rows,
                    min_col=1,
                    max_col=len(headers)
                ):
                    for cell in row:
                        cell.border = thin_border
        
        except Exception as e:
            log.warning(f"Error applying worksheet formatting: {e}")
    
    def _create_metadata_sheet(self, workbook, metadata: Dict[str, Any]) -> None:
        """Create a separate metadata worksheet."""
        try:
            metadata_sheet = workbook.create_sheet(title="Export Metadata")
            
            # Write metadata in a nice format
            current_row = 1
            
            for key, value in metadata.items():
                display_key = key.replace('_', ' ').title()
                metadata_sheet.cell(row=current_row, column=1, value=display_key)
                metadata_sheet.cell(row=current_row, column=2, value=str(value))
                current_row += 1
            
            # Style the metadata sheet
            for row in metadata_sheet.iter_rows(min_row=1, max_row=current_row-1, min_col=1, max_col=2):
                row[0].font = Font(bold=True)
                row[0].alignment = Alignment(horizontal='right')
        
        except Exception as e:
            log.warning(f"Error creating metadata sheet: {e}")
    
    def _add_charts(self, worksheet, data: List[Dict[str, Any]], options: Dict[str, Any]) -> None:
        """Add charts to the worksheet based on data."""
        try:
            # This is a simplified chart implementation
            # In a full implementation, you'd want more sophisticated chart logic
            
            chart_types = options.get('chart_types', ['bar'])
            
            # Find numeric columns for charting
            numeric_columns = []
            if data:
                for key, value in data[0].items():
                    if isinstance(value, (int, float)) and not isinstance(value, bool):
                        numeric_columns.append(key)
            
            if len(numeric_columns) > 0 and len(data) > 1:
                # Create a simple bar chart with first numeric column
                if 'bar' in chart_types:
                    chart = BarChart()
                    chart.title = f"Chart: {numeric_columns[0]}"
                    
                    # Add chart (simplified - real implementation would be more complex)
                    chart_pos = f"{chr(ord('A') + len(data[0]) + 2)}2"
                    worksheet.add_chart(chart, chart_pos)
        
        except Exception as e:
            log.warning(f"Error adding charts: {e}")
    
    def export_to_file(self, 
                       data: List[Dict[str, Any]],
                       filepath: str,
                       metadata: Optional[Dict[str, Any]] = None,
                       options: Optional[Dict[str, Any]] = None) -> None:
        """
        Export data directly to an Excel file.
        
        Args:
            data: Data to export
            filepath: File path to write to
            metadata: Export metadata
            options: Export options
        """
        try:
            excel_content = self.export(data, filepath, metadata, options)
            
            with open(filepath, 'wb') as f:
                f.write(excel_content)
            
            log.info(f"Excel file written successfully: {filepath}")
            
        except Exception as e:
            log.error(f"Error writing Excel file {filepath}: {e}")
            raise
    
    def get_sample_options(self) -> Dict[str, Any]:
        """Get sample configuration options."""
        return {
            'sheet_name': {
                'value': 'Data',
                'description': 'Name of the main worksheet'
            },
            'include_metadata': {
                'value': True,
                'description': 'Include export metadata at the top of the sheet'
            },
            'auto_filter': {
                'value': True,
                'description': 'Add auto-filter to header row'
            },
            'freeze_headers': {
                'value': True,
                'description': 'Freeze the header row for easier scrolling'
            },
            'header_style': {
                'value': {
                    'font': {'bold': True, 'color': 'FFFFFF'},
                    'fill': {'color': '4F81BD', 'type': 'solid'}
                },
                'description': 'Styling options for header cells'
            },
            'add_charts': {
                'value': False,
                'description': 'Automatically add charts based on numeric data'
            },
            'metadata_sheet': {
                'value': False,
                'description': 'Create a separate sheet for export metadata'
            }
        }